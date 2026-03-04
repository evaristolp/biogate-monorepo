from __future__ import annotations

from typing import TYPE_CHECKING, Any, Dict, List, Optional, Sequence, Tuple

# NOTE:
# We intentionally avoid importing openpyxl (and therefore numpy) at module
# import time. In some environments a broken numpy/Accelerate stack can
# segfault during import, which would take down pytest collection and any
# code path that merely imports this module. Instead we import openpyxl
# lazily inside extract_from_excel() the first time it's actually used.
if TYPE_CHECKING:  # pragma: no cover - import only for type checking
    from openpyxl.worksheet.worksheet import Worksheet

from backend.ingestion.base import (
    ExtractionMethod,
    ExtractionResult,
    ExtractedVendor,
)


VENDOR_NAME_KEYWORDS: List[str] = [
    "vendor",
    "supplier",
    "manufacturer",
    "company",
    "name",
    "firm",
    "provider",
]

COUNTRY_KEYWORDS: List[str] = [
    "country",
    "nation",
    "location",
    "origin",
    "headquarters",
    "hq",
]

EQUIPMENT_KEYWORDS: List[str] = [
    "product",
    "equipment",
    "item",
    "description",
    "service",
    "category",
    "type",
]


def _cell_value(cell: Any) -> Any:
    if hasattr(cell, "value"):
        return cell.value
    return cell


def _normalize_header_value(value: Any) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    return value.strip().lower()


def detect_columns(
    rows: Sequence[Sequence[Any]],
) -> Tuple[Optional[int], Dict[str, int]]:
    """
    Scan up to the first 10 rows to detect header row and column indices.

    Returns a tuple of (header_row_index, col_map) where header_row_index is
    the index into `rows` (0-based) and col_map maps the logical keys
    'name', 'country', 'equipment' to column indices (0-based).
    """
    max_rows_to_scan = min(10, len(rows))

    for row_idx in range(max_rows_to_scan):
        row = rows[row_idx]
        col_map: Dict[str, int] = {}

        for col_idx, cell in enumerate(row):
            header_text = _normalize_header_value(_cell_value(cell))
            if not header_text:
                continue

            if "name" not in col_map and any(
                kw in header_text for kw in VENDOR_NAME_KEYWORDS
            ):
                col_map["name"] = col_idx

            if "country" not in col_map and any(
                kw in header_text for kw in COUNTRY_KEYWORDS
            ):
                col_map["country"] = col_idx

            if "equipment" not in col_map and any(
                kw in header_text for kw in EQUIPMENT_KEYWORDS
            ):
                col_map["equipment"] = col_idx

        if "name" in col_map:
            return row_idx, col_map

    return None, {}


def extract_vendor_from_row(
    row: Sequence[Any],
    col_map: Dict[str, int],
) -> Optional[ExtractedVendor]:
    """
    Extract a single vendor from a worksheet row using the detected columns.

    Returns None if the row does not contain a valid vendor (e.g. empty or too short name).
    """
    name_idx = col_map.get("name")
    if name_idx is None or name_idx >= len(row):
        return None

    raw_name = _cell_value(row[name_idx])
    if raw_name is None:
        return None

    if not isinstance(raw_name, str):
        raw_name = str(raw_name)

    raw_name = raw_name.strip()
    if len(raw_name) < 2:
        return None

    country_hint: Optional[str] = None
    equipment_hint: Optional[str] = None

    country_idx = col_map.get("country")
    if country_idx is not None and country_idx < len(row):
        value = _cell_value(row[country_idx])
        if value is not None:
            country_hint = str(value).strip() or None

    equipment_idx = col_map.get("equipment")
    if equipment_idx is not None and equipment_idx < len(row):
        value = _cell_value(row[equipment_idx])
        if value is not None:
            equipment_hint = str(value).strip() or None

    return ExtractedVendor(
        raw_name=raw_name,
        country_hint=country_hint,
        equipment_type_hint=equipment_hint,
    )


def _extract_from_sheet(sheet: "Worksheet") -> List[ExtractedVendor]:
    rows = list(sheet.iter_rows(values_only=False))
    if not rows:
        return []

    header_row_rel_idx, col_map = detect_columns(rows[:10])
    if header_row_rel_idx is None:
        return []

    header_row_abs_idx = header_row_rel_idx  # still 0-based into `rows`
    vendors: List[ExtractedVendor] = []

    for row_offset, row in enumerate(rows[header_row_abs_idx + 1 :], start=header_row_abs_idx + 1):
        vendor = extract_vendor_from_row(row, col_map)
        if not vendor:
            continue

        # Row numbers in Excel are 1-based, and `rows` is 0-based.
        excel_row_number = row_offset + 1
        vendor.extraction_confidence = 0.85
        vendor.source_context = f"Sheet '{sheet.title}', Row {excel_row_number}"
        vendors.append(vendor)

    return vendors


def extract_from_excel(file_path: str) -> ExtractionResult:
    """
    Extract vendor records from an Excel (.xlsx) file.
    """
    result = ExtractionResult(extraction_method=ExtractionMethod.EXCEL_PARSER)

    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - defensive
        result.errors.append(f"openpyxl is not available: {exc}")
        return result

    try:
        wb = load_workbook(filename=file_path, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive
        result.errors.append(f"Failed to open Excel file: {exc}")
        return result

    all_vendors: List[ExtractedVendor] = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_vendors = _extract_from_sheet(sheet)
        all_vendors.extend(sheet_vendors)

    result.vendors = all_vendors

    if all_vendors:
        # Overall extraction confidence mirrors per-record default.
        result.extraction_confidence = 0.85

    return result

